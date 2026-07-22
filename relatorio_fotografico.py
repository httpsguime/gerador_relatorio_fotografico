import argparse
import io
import os
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl
from PIL import Image, ImageOps
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXTENSOES_IMG = (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp")

# Nome de aba preferido na planilha de atendimentos, se existir.
ABA_PREFERIDA = "Atendimentos"

# Nomes de cabeçalho aceitos para cada campo (comparados sem acento e minúsculos).
# Ajuste aqui se algum contrato usar nomes de coluna diferentes.
ALIASES_COLUNAS = {
    "numero": {"atendimento", "numero atendimento", "no atendimento", "codigo atendimento"},
    "logradouro": {"logradouro", "endereco", "rua"},
    "numero_casa": {"no", "n", "num", "numero casa", "numero do imovel"},
    "bairro": {"bairro"},
    "cidade": {"cidade", "municipio"},
}

# Trechos "burocráticos" do endereço (vindos de geocodificação) que devem ser
# descartados, mantendo só rua, número, bairro e cidade.
_CEP_RE = re.compile(r"^\d{5}-?\d{3}$")
_PARTES_DESCARTAVEIS = {"brasil", "brazil", "rio de janeiro", "região sudeste"}
_PREFIXOS_DESCARTAVEIS = (
    "região geográfica imediata",
    "região metropolitana",
    "região geográfica intermediária",
)


def simplificar_endereco(endereco):
    """Reduz um endereço vindo de geocodificação a rua, número, bairro e cidade."""
    partes_originais = [p.strip() for p in endereco.split(",") if p.strip()]

    partes = []
    for p in partes_originais:
        p_lower = p.lower()
        if _CEP_RE.match(p):
            continue
        if p_lower in _PARTES_DESCARTAVEIS:
            continue
        if p_lower.startswith(_PREFIXOS_DESCARTAVEIS):
            continue
        partes.append(p)

    # remove duplicatas mantendo a primeira ocorrência (ex.: cidade repetida no fim)
    vistas = set()
    sem_duplicatas = []
    for p in partes:
        chave = p.lower()
        if chave not in vistas:
            vistas.add(chave)
            sem_duplicatas.append(p)
    partes = sem_duplicatas

    # move o número da casa (parte só com dígitos) para logo após a rua
    numeros = [p for p in partes[1:] if p.isdigit()]
    if numeros and len(partes) > 1:
        resto = [p for p in partes[1:] if not p.isdigit()]
        partes = [partes[0]] + numeros + resto

    return ", ".join(partes) if partes else endereco.strip()

# Compressão das fotos no PDF (evita arquivo gigante).
MAX_DIMENSAO_PX = 1600   # maior lado da imagem, em pixels
JPEG_QUALIDADE = 75      # qualidade JPEG (0-100)

# ---------------------------------------------------------------------------
# Capas (padrão dos arquivos "Capas Fotográficos.pdf" e "Capas Fotográficos 2.pdf")
# ---------------------------------------------------------------------------
# Cada contrato tem seu próprio brasão. "nome" é usado no título da capa
# (igual ao texto do PDF de referência); "label" é o texto mostrado na tela.
# "rotulo_data" só é preciso informar quando o contrato foge do padrão "PERÍODO"
# (ex.: Porto Real usa "DATA DE EXECUÇÃO"). "layout" só é preciso quando o
# contrato tem uma capa totalmente diferente do padrão (ex.: Casimiro de Abreu).
CONTRATOS = {
    "campos": {
        "nome": "CAMPOS DOS GOYTACAZES",
        "label": "Campos dos Goytacazes",
        "brasao": "brasao_campos.png",
    },
    "sao_goncalo": {
        "nome": "SÃO GONÇALO",
        "label": "São Gonçalo",
        "brasao": "brasao_sao_goncalo.png",
    },
    "saquarema": {
        "nome": "SAQUAREMA",
        "label": "Saquarema",
        "brasao": "brasao_saquarema.png",
    },
    "belford_roxo": {
        "nome": "BELFORD ROXO",
        "label": "Belford Roxo",
        "brasao": "brasao_belford_roxo.png",
    },
    "buri": {
        "nome": "BURI",
        "label": "Buri",
        "brasao": "brasao_buri.png",
    },
    "cabo_santo_agostinho": {
        "nome": "CABO DE SANTO AGOSTINHO",
        "label": "Cabo de Santo Agostinho",
        "brasao": "brasao_cabo_santo_agostinho.png",
    },
    "cabo_frio": {
        "nome": "CABO FRIO",
        "label": "Cabo Frio",
        "brasao": "brasao_cabo_frio.png",
    },
    "casimiro_abreu": {
        "nome": "CASIMIRO DE ABREU",
        "label": "Casimiro de Abreu",
        "brasao": "logo_casimiro_abreu.png",
        "layout": "casimiro",
    },
    "levy_gasparian": {
        "nome": "LEVY GASPARIAN",
        "label": "Levy Gasparian",
        "brasao": "brasao_levy_gasparian.png",
    },
    "macae": {
        "nome": "MACAÉ",
        "label": "Macaé",
        "brasao": "brasao_macae.png",
    },
    "mage": {
        "nome": "MAGÉ",
        "label": "Magé",
        "brasao": "brasao_mage.png",
    },
    "paty_do_alferes": {
        "nome": "PATY DO ALFERES",
        "label": "Paty do Alferes",
        "brasao": "brasao_paty_do_alferes.png",
    },
    "porto_real": {
        "nome": "PORTO REAL",
        "label": "Porto Real",
        "brasao": "brasao_porto_real.png",
        "rotulo_data": "DATA DE EXECUÇÃO",
    },
}

# "conector" é o trecho que segue "SERVIÇO" no título (a maioria leva "DE ...",
# mas "MODERNIZAÇÃO" no PDF de referência aparece sem o "DE").
TIPOS_RELATORIO = {
    "implantacao": {"nome": "IMPLANTAÇÃO", "conector": "DE IMPLANTAÇÃO"},
    "manutencao": {"nome": "MANUTENÇÃO", "conector": "DE MANUTENÇÃO"},
    "modernizacao": {"nome": "MODERNIZAÇÃO", "conector": "MODERNIZAÇÃO"},
}

# Combinações contrato+tipo que fogem do padrão "PERÍODO: DE ... À ..." e usam
# "MÊS / ANO" na capa (Campos/Manutenção e Cabo Frio/Manutenção, no PDF de referência).
EXCECOES_MES_ANO = {
    ("campos", "manutencao"),
    ("cabo_frio", "manutencao"),
}

# Combinações contrato+tipo cujo título encurta para "SERVIÇO {CONECTOR} DE
# {CIDADE}", sem o "DO PARQUE DE ILUMINAÇÃO PÚBLICA" do padrão normal.
EXCECOES_TITULO_CURTO = {
    ("campos", "manutencao"),
}

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def caminho_asset(nome_arquivo):
    """Caminho de um arquivo em assets/, tanto rodando como script quanto
    empacotado em .exe (PyInstaller extrai os dados para sys._MEIPASS)."""
    base = getattr(sys, "_MEIPASS", BASE_DIR)
    return os.path.join(base, "assets", nome_arquivo)


def usa_layout_casimiro(contrato_id):
    """Casimiro de Abreu tem uma capa totalmente diferente do padrão: logo da
    prefeitura (não um brasão), título genérico sem o nome da cidade, fonte
    serifada e sem campo de data."""
    return CONTRATOS[contrato_id].get("layout") == "casimiro"


def usa_mes_ano(contrato_id, tipo_id):
    """True para as combinações contrato+tipo que usam 'MÊS / ANO' na capa em
    vez de 'PERÍODO: DE ... À ...' (ver EXCECOES_MES_ANO)."""
    return (contrato_id, tipo_id) in EXCECOES_MES_ANO


def rotulo_periodo(contrato_id):
    """Rótulo usado antes da data no formato período (normalmente 'PERÍODO',
    mas alguns contratos, como Porto Real, usam outro texto)."""
    return CONTRATOS[contrato_id].get("rotulo_data", "PERÍODO")


def montar_titulo_capa(contrato_id, tipo_id):
    """Monta o título da capa seguindo o padrão dos PDFs de referência."""
    tipo_nome = TIPOS_RELATORIO[tipo_id]["nome"]
    if usa_layout_casimiro(contrato_id):
        return f"RELATÓRIO FOTOGRÁFICO DE {tipo_nome}"

    cidade = CONTRATOS[contrato_id]["nome"]
    conector = TIPOS_RELATORIO[tipo_id]["conector"]
    if (contrato_id, tipo_id) in EXCECOES_TITULO_CURTO:
        return f"RELATÓRIO FOTOGRÁFICO REFERENTE AO SERVIÇO {conector} DE {cidade}"
    return (
        f"RELATÓRIO FOTOGRÁFICO REFERENTE AO SERVIÇO {conector} "
        f"DO PARQUE DE ILUMINAÇÃO PÚBLICA DE {cidade}"
    )


def _normalizar_cabecalho(valor):
    """Remove acentos/caixa de um valor de cabeçalho para comparação flexível."""
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto.strip().lower()


def _mapear_colunas(linha_cabecalho):
    """Descobre em quais índices da linha estão os campos que precisamos."""
    mapa = {}
    for indice, valor in enumerate(linha_cabecalho):
        norm = _normalizar_cabecalho(valor)
        for campo, aliases in ALIASES_COLUNAS.items():
            if campo not in mapa and norm in aliases:
                mapa[campo] = indice
    return mapa


def _texto_celula(linha, mapa, campo):
    """Lê o valor de uma célula (por nome de campo) já convertido para texto limpo."""
    if campo not in mapa:
        return ""
    valor = linha[mapa[campo]]
    if valor is None or valor == "":
        return ""
    texto = str(valor).strip()
    # openpyxl entrega números inteiros como float (ex.: 80 -> "80.0"); corrige.
    if texto.endswith(".0") and texto[:-2].lstrip("-").isdigit():
        texto = texto[:-2]
    return texto


def ler_atendimentos(caminho_planilha):
    """Lê a planilha de atendimentos (.xlsx) e retorna uma lista de (numero, endereco)."""
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True, read_only=True)
    try:
        aba = wb[ABA_PREFERIDA] if ABA_PREFERIDA in wb.sheetnames else wb.worksheets[0]
        linhas = list(aba.iter_rows(values_only=True))
    finally:
        wb.close()

    mapa = None
    inicio_dados = None
    for indice, linha in enumerate(linhas):
        candidato = _mapear_colunas(linha)
        if "numero" in candidato and "logradouro" in candidato:
            mapa = candidato
            inicio_dados = indice + 1
            break

    if mapa is None:
        sys.exit(
            "Não foi possível encontrar as colunas de atendimento/logradouro "
            f"na planilha: {caminho_planilha}"
        )

    atendimentos = []
    for linha in linhas[inicio_dados:]:
        if not linha:
            continue
        numero = _texto_celula(linha, mapa, "numero")
        if not numero:
            continue

        partes_endereco = [
            _texto_celula(linha, mapa, campo)
            for campo in ("logradouro", "numero_casa", "bairro", "cidade")
        ]
        endereco = ", ".join(p for p in partes_endereco if p)
        atendimentos.append((numero, endereco))






    return atendimentos


def listar_fotos(pasta):
    """Retorna a lista (ordenada) de caminhos de imagem dentro da pasta, ou [] se não existir/vazia."""
    if not os.path.isdir(pasta):
        return []
    return [
        os.path.join(pasta, nome)
        for nome in sorted(os.listdir(pasta))
        if nome.lower().endswith(EXTENSOES_IMG)
    ]


def listar_pastas_sem_atendimento(fotos_dir, numeros_atendimentos):
    """Pastas dentro de fotos_dir que têm fotos mas não correspondem a nenhum
    atendimento da planilha (fotos que, portanto, não entram no relatório)."""
    orfas = []
    if not os.path.isdir(fotos_dir):
        return orfas
    for nome in sorted(os.listdir(fotos_dir)):
        caminho = os.path.join(fotos_dir, nome)
        if not os.path.isdir(caminho) or nome in numeros_atendimentos:
            continue
        fotos = listar_fotos(caminho)
        if fotos:
            orfas.append((nome, [os.path.basename(f) for f in fotos]))
    return orfas


def _quebrar_texto(c, texto, fonte, tamanho, largura_max):
    """Quebra o texto em várias linhas para caber na largura disponível."""
    palavras = texto.split()
    linhas = []
    atual = ""
    for palavra in palavras:
        teste = f"{atual} {palavra}".strip()
        if c.stringWidth(teste, fonte, tamanho) <= largura_max:
            atual = teste
        else:
            if atual:
                linhas.append(atual)
            atual = palavra
    if atual:
        linhas.append(atual)
    return linhas or [texto]


def desenhar_pagina(c, largura, altura, indice, endereco, caminho_foto):
    """Desenha uma página do relatório."""
    margem = 2 * cm

    # Cabeçalho: número sequencial (não é o número real do atendimento)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(margem, altura - margem, f"Atendimento {indice}")

    # Endereço (simplificado e com quebra de linha se ainda for muito longo)
    c.setFont("Helvetica", 12)
    endereco_curto = simplificar_endereco(endereco) if endereco else ""
    texto_endereco = f"Endereço: {endereco_curto}" if endereco_curto else "Endereço: (não informado)"
    largura_util = largura - 2 * margem
    y_endereco = altura - margem - 1 * cm
    for linha in _quebrar_texto(c, texto_endereco, "Helvetica", 12, largura_util):
        c.drawString(margem, y_endereco, linha)
        y_endereco -= 0.6 * cm

    # Área da foto
    topo_foto = y_endereco - 0.5 * cm
    base_foto = margem
    area_larg = largura - 2 * margem
    area_alt = topo_foto - base_foto

    if caminho_foto:
        try:
            img = Image.open(caminho_foto)
            img = ImageOps.exif_transpose(img)  # corrige orientação pelo EXIF
            img = img.convert("RGB")

            # Reduz a resolução para não gerar um PDF enorme.
            img.thumbnail((MAX_DIMENSAO_PX, MAX_DIMENSAO_PX), Image.LANCZOS)

            # Recomprime como JPEG em memória.
            buffer = io.BytesIO()
            img.save(buffer, format="JPEG", quality=JPEG_QUALIDADE, optimize=True)
            buffer.seek(0)

            iw, ih = img.size
            escala = min(area_larg / iw, area_alt / ih)
            nova_larg = iw * escala
            nova_alt = ih * escala
            x = margem + (area_larg - nova_larg) / 2
            y = base_foto + (area_alt - nova_alt) / 2
            c.drawImage(
                ImageReader(buffer),
                x, y, width=nova_larg, height=nova_alt,
                preserveAspectRatio=True, anchor="c",
            )
        except Exception as e:
            c.setFont("Helvetica-Oblique", 11)
            c.drawString(margem, topo_foto - 1 * cm, f"[Erro ao carregar a foto: {e}]")
    else:
        c.setFont("Helvetica-Oblique", 11)
        c.drawString(margem, topo_foto - 1 * cm,
                     "[Nenhuma foto encontrada para este atendimento]")

    c.showPage()


def _desenhar_logo_hashimoto(c, largura):
    """Desenha a logo da Hashimoto fixa no rodapé da capa, centralizada."""
    centro_x = largura / 2
    logo = ImageReader(caminho_asset("logo_hashimoto.png"))
    lw, lh = logo.getSize()
    max_larg_logo, max_alt_logo = 7 * cm, 5 * cm
    escala_logo = min(max_larg_logo / lw, max_alt_logo / lh)
    l_larg, l_alt = lw * escala_logo, lh * escala_logo
    base_logo = 1.6 * cm
    c.drawImage(
        logo, centro_x - l_larg / 2, base_logo, width=l_larg, height=l_alt,
        preserveAspectRatio=True, anchor="c", mask="auto",
    )


def desenhar_capa(c, largura, altura, contrato_id, tipo_id, texto_data):
    """Desenha a página de capa: brasão do contrato, título do tipo de
    serviço, data/período informados e a logo da Hashimoto, seguindo o
    padrão dos PDFs 'Capas Fotográficos'. Casimiro de Abreu usa um layout
    próprio (ver desenhar_capa_casimiro)."""
    if usa_layout_casimiro(contrato_id):
        desenhar_capa_casimiro(c, largura, altura, tipo_id)
        return

    margem = 2.2 * cm
    largura_util = largura - 2 * margem
    centro_x = largura / 2

    # Brasão do contrato
    brasao = ImageReader(caminho_asset(CONTRATOS[contrato_id]["brasao"]))
    iw, ih = brasao.getSize()
    max_larg_brasao, max_alt_brasao = 7.5 * cm, 9 * cm
    escala = min(max_larg_brasao / iw, max_alt_brasao / ih)
    b_larg, b_alt = iw * escala, ih * escala
    topo_brasao = altura - 1.6 * cm
    c.drawImage(
        brasao, centro_x - b_larg / 2, topo_brasao - b_alt, width=b_larg, height=b_alt,
        preserveAspectRatio=True, anchor="c", mask="auto",
    )

    # Título (nome do contrato + tipo de serviço)
    fonte_titulo, tamanho_titulo = "Helvetica-Bold", 20
    c.setFont(fonte_titulo, tamanho_titulo)
    titulo = montar_titulo_capa(contrato_id, tipo_id)
    y = topo_brasao - b_alt - 1.3 * cm
    for linha in _quebrar_texto(c, titulo, fonte_titulo, tamanho_titulo, largura_util):
        c.drawCentredString(centro_x, y, linha)
        y -= tamanho_titulo * 1.3

    # Data / período informado
    fonte_data, tamanho_data = "Helvetica-Bold", 15
    c.setFont(fonte_data, tamanho_data)
    y -= 1.2 * cm
    c.drawCentredString(centro_x, y, texto_data)

    _desenhar_logo_hashimoto(c, largura)
    c.showPage()


def desenhar_capa_casimiro(c, largura, altura, tipo_id):
    """Layout específico de Casimiro de Abreu: logo da prefeitura (não um
    brasão), título genérico em fonte serifada (sem o nome da cidade) e sem
    campo de data, seguindo o padrão do PDF de referência."""
    centro_x = largura / 2
    margem = 2.2 * cm
    largura_util = largura - 2 * margem

    logo = ImageReader(caminho_asset(CONTRATOS["casimiro_abreu"]["brasao"]))
    lw, lh = logo.getSize()
    max_larg_logo, max_alt_logo = 10 * cm, 3.6 * cm
    escala = min(max_larg_logo / lw, max_alt_logo / lh)
    l_larg, l_alt = lw * escala, lh * escala
    topo_logo = altura - 1.6 * cm
    c.drawImage(
        logo, centro_x - l_larg / 2, topo_logo - l_alt, width=l_larg, height=l_alt,
        preserveAspectRatio=True, anchor="c", mask="auto",
    )

    fonte_titulo, tamanho_titulo = "Times-Bold", 34
    c.setFont(fonte_titulo, tamanho_titulo)
    titulo = montar_titulo_capa("casimiro_abreu", tipo_id)
    y = topo_logo - l_alt - 2 * cm
    for linha in _quebrar_texto(c, titulo, fonte_titulo, tamanho_titulo, largura_util):
        c.drawCentredString(centro_x, y, linha)
        y -= tamanho_titulo * 1.25

    _desenhar_logo_hashimoto(c, largura)
    c.showPage()


def gerar_nome_saida(caminho_planilha, pasta_fotos, pasta_saida=None):
    """Gera o caminho do PDF de saída. Por padrão, salva ao lado da planilha;
    se `pasta_saida` for informada, salva nela em vez disso."""
    nome_base = os.path.basename(os.path.normpath(pasta_fotos)) or "relatorio"
    diretorio = pasta_saida or os.path.dirname(os.path.abspath(caminho_planilha))
    return os.path.join(diretorio, f"relatorio_fotografico_{nome_base}.pdf")


def gerar_nome_log(saida_pdf):
    """Gera o caminho do arquivo de log (.txt) a partir do caminho do PDF de saída."""
    base, _ext = os.path.splitext(saida_pdf)
    return f"{base}_log.txt"


@dataclass
class ResultadoRelatorio:
    total: int
    com_foto: int
    sem_foto: int
    atendimentos_sem_foto: list = field(default_factory=list)   # [(numero, endereco), ...]
    pastas_sem_atendimento: list = field(default_factory=list)  # [(nome_pasta, [arquivos]), ...]
    log_texto: str = ""
    log_path: str = ""


def montar_log(planilha_path, fotos_dir, saida_pdf, resultado):
    """Monta o texto do log com os erros encontrados na execução."""
    L = []
    L.append("Relatório Fotográfico - Log de Erros")
    L.append(f"Planilha:      {planilha_path}")
    L.append(f"Pack de fotos: {fotos_dir}")
    L.append(f"PDF gerado:    {saida_pdf}")
    L.append(f"Gerado em:     {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    L.append("")
    L.append("RESUMO")
    L.append(f"  Atendimentos na planilha:........................ {resultado.total}")
    L.append(f"  Atendimentos com foto:............................ {resultado.com_foto}")
    L.append(f"  Atendimentos SEM foto:............................ {resultado.sem_foto}")
    L.append(f"  Pastas de fotos sem atendimento correspondente:.. {len(resultado.pastas_sem_atendimento)}")
    L.append("")

    L.append(f"ATENDIMENTOS SEM FOTO ({len(resultado.atendimentos_sem_foto)})")
    if resultado.atendimentos_sem_foto:
        for numero, endereco in resultado.atendimentos_sem_foto:
            L.append(f"  - Atendimento {numero}: {endereco or '(endereço não informado)'}")
    else:
        L.append("  Nenhum. Todos os atendimentos têm foto.")
    L.append("")

    L.append(f"PASTAS DE FOTOS SEM ATENDIMENTO CORRESPONDENTE ({len(resultado.pastas_sem_atendimento)})")
    if resultado.pastas_sem_atendimento:
        for nome_pasta, arquivos in resultado.pastas_sem_atendimento:
            L.append(f"  - Pasta '{nome_pasta}' ({len(arquivos)} foto(s)): {', '.join(arquivos)}")
    else:
        L.append("  Nenhuma. Toda pasta do pack de fotos corresponde a um atendimento da planilha.")

    return "\n".join(L)


def gerar_relatorio(planilha_path, fotos_dir, saida_pdf, capa=None, ao_avancar=None):
    """Gera o PDF do relatório e um log (.txt) com os erros encontrados ao lado dele.
    `capa`, se informado, é um dict {"contrato_id", "tipo_id", "texto_data"} usado
    para desenhar a página de capa antes dos atendimentos.
    `ao_avancar(indice, total)` é chamado a cada página."""
    atendimentos = ler_atendimentos(planilha_path)
    total = len(atendimentos)

    c = canvas.Canvas(saida_pdf, pagesize=A4)
    largura, altura = A4

    if capa:
        desenhar_capa(c, largura, altura, capa["contrato_id"], capa["tipo_id"], capa["texto_data"])

    com_foto = 0
    atendimentos_sem_foto = []
    numeros_atendimentos = {numero for numero, _ in atendimentos}

    for indice, (numero, endereco) in enumerate(atendimentos, start=1):
        pasta = os.path.join(fotos_dir, numero)
        fotos = listar_fotos(pasta)
        foto = fotos[0] if fotos else None
        if foto:
            com_foto += 1
        else:
            atendimentos_sem_foto.append((numero, endereco))
        desenhar_pagina(c, largura, altura, indice, endereco, foto)
        if ao_avancar:
            ao_avancar(indice, total)

    c.save()

    pastas_sem_atendimento = listar_pastas_sem_atendimento(fotos_dir, numeros_atendimentos)

    resultado = ResultadoRelatorio(
        total=total,
        com_foto=com_foto,
        sem_foto=len(atendimentos_sem_foto),
        atendimentos_sem_foto=atendimentos_sem_foto,
        pastas_sem_atendimento=pastas_sem_atendimento,
    )
    resultado.log_texto = montar_log(planilha_path, fotos_dir, saida_pdf, resultado)
    resultado.log_path = gerar_nome_log(saida_pdf)
    with open(resultado.log_path, "w", encoding="utf-8") as f:
        f.write(resultado.log_texto)

    return resultado


def executar_cli(planilha_path, fotos_dir, saida_pdf, capa=None):
    """Gera o relatório direto no terminal, sem interface gráfica."""
    if not os.path.exists(planilha_path):
        sys.exit(f"Planilha não encontrada: {planilha_path}")
    if not os.path.isdir(fotos_dir):
        sys.exit(f"Pasta de fotos não encontrada: {fotos_dir}")

    saida_pdf = saida_pdf or gerar_nome_saida(planilha_path, fotos_dir)
    print(f"Planilha: {planilha_path}")
    print(f"Pack de fotos: {fotos_dir}")

    resultado = gerar_relatorio(planilha_path, fotos_dir, saida_pdf, capa=capa)

    print(f"{resultado.total} atendimentos lidos da planilha.")
    print(f"Relatório gerado: {saida_pdf}")
    print(f"  Páginas com foto: {resultado.com_foto}")
    print(f"  Páginas sem foto: {resultado.sem_foto}")
    print()
    print(resultado.log_texto)
    print()
    print(f"Log completo salvo em: {resultado.log_path}")


# ---------------------------------------------------------------------------
# Interface gráfica (janela guiada passo a passo)
# ---------------------------------------------------------------------------
class JanelaRelatorio:
    """Janela única que guia o usuário: escolher planilha, pasta de fotos e gerar o PDF."""

    # Paleta de cores da interface.
    COR_FUNDO = "#f4f5f9"
    COR_CARTAO = "#ffffff"
    COR_BORDA = "#e2e4ea"
    COR_TEXTO = "#1f2933"
    COR_TEXTO_SEC = "#6b7280"
    COR_ACCENT = "#4f46e5"
    COR_ACCENT_HOVER = "#4338ca"
    COR_SUCESSO = "#15803d"
    COR_ERRO = "#dc2626"
    LARGURA_CONTEUDO = 520

    def __init__(self, root, planilha_inicial=None, fotos_inicial=None):
        import tkinter as tk
        from tkinter import ttk

        self._tk = tk
        self._ttk = ttk
        self.root = root
        root.title("Relatório Fotográfico")
        root.resizable(False, False)
        root.configure(background=self.COR_FUNDO)

        self._configurar_estilos(ttk)

        self.planilha_var = tk.StringVar(value=planilha_inicial or "")
        self.fotos_var = tk.StringVar(value=fotos_inicial or "")
        self.saida_var = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="Selecione a planilha e a pasta de fotos para começar.")

        # Passo 4: tipo de serviço, contrato e data da capa.
        self.tipo_var = tk.StringVar(value="implantacao")
        self._labels_contrato = [info["label"] for info in CONTRATOS.values()]
        self._contrato_por_label = {info["label"]: cid for cid, info in CONTRATOS.items()}
        self.contrato_label_var = tk.StringVar(value=self._labels_contrato[0])
        agora = datetime.now()
        self.mes_var = tk.StringVar(value=MESES_PT[agora.month - 1])
        self.ano_var = tk.StringVar(value=str(agora.year))
        self.periodo_de_var = tk.StringVar(value="")
        self.periodo_ate_var = tk.StringVar(value="")

        externo = ttk.Frame(root, style="Fundo.TFrame", padding=24)
        externo.grid(sticky="nsew")

        ttk.Label(externo, text="Relatório Fotográfico", style="Titulo.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(
            externo,
            text="Gera um PDF cruzando a planilha de atendimentos com as fotos do pack.",
            style="Subtitulo.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(2, 18))

        self._criar_passo(
            externo, row=2, numero="①",
            titulo="Selecione o arquivo com a lista de atendimentos (Excel)",
            variavel=self.planilha_var,
            texto_botao="Selecionar arquivo...",
            comando=self.selecionar_planilha,
        )
        self._criar_passo(
            externo, row=3, numero="②",
            titulo="Selecione a pasta com o pack de fotos",
            variavel=self.fotos_var,
            texto_botao="Selecionar pasta...",
            comando=self.selecionar_fotos,
        )
        self._criar_passo(
            externo, row=4, numero="③",
            titulo="Pasta onde salvar o PDF (opcional — por padrão, salva ao lado da planilha)",
            variavel=self.saida_var,
            texto_botao="Selecionar pasta...",
            comando=self.selecionar_saida,
        )

        self._criar_passo_capa(externo, row=5)

        rodape = ttk.Frame(externo, style="Fundo.TFrame")
        rodape.grid(row=6, column=0, sticky="we", pady=(6, 0))
        rodape.columnconfigure(0, weight=1)

        self.status_label = ttk.Label(
            rodape, textvariable=self.status_var, style="Status.TLabel",
            wraplength=self.LARGURA_CONTEUDO,
        )
        self.status_label.grid(row=0, column=0, sticky="w")

        self.progress = ttk.Progressbar(
            rodape, length=self.LARGURA_CONTEUDO, mode="determinate",
            style="Accent.Horizontal.TProgressbar",
        )
        self.progress.grid(row=1, column=0, sticky="we", pady=(10, 16))

        self.gerar_btn = ttk.Button(
            rodape, text="Gerar Relatório", command=self.gerar, state="disabled",
            style="Accent.TButton", cursor="hand2",
        )
        self.gerar_btn.grid(row=2, column=0, sticky="we")

        self._atualizar_estado()
        self._centralizar_janela()

    def _configurar_estilos(self, ttk):
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("Fundo.TFrame", background=self.COR_FUNDO)
        style.configure("Cartao.TFrame", background=self.COR_CARTAO)
        style.configure(
            "Cartao.TLabel", background=self.COR_CARTAO, foreground=self.COR_TEXTO,
            font=("Segoe UI", 10),
        )
        style.configure(
            "CartaoTitulo.TLabel", background=self.COR_CARTAO, foreground=self.COR_TEXTO,
            font=("Segoe UI", 10, "bold"),
        )
        style.configure(
            "Numero.TLabel", background=self.COR_CARTAO, foreground=self.COR_ACCENT,
            font=("Segoe UI", 13, "bold"),
        )
        style.configure(
            "Titulo.TLabel", background=self.COR_FUNDO, foreground=self.COR_TEXTO,
            font=("Segoe UI", 18, "bold"),
        )
        style.configure(
            "Subtitulo.TLabel", background=self.COR_FUNDO, foreground=self.COR_TEXTO_SEC,
            font=("Segoe UI", 10),
        )
        style.configure(
            "Status.TLabel", background=self.COR_FUNDO, foreground=self.COR_TEXTO_SEC,
            font=("Segoe UI", 10),
        )

        style.configure(
            "Cartao.TEntry", fieldbackground="#f9fafb", foreground=self.COR_TEXTO,
            bordercolor=self.COR_BORDA, lightcolor=self.COR_BORDA, darkcolor=self.COR_BORDA,
            padding=6,
        )

        style.configure("Secundario.TButton", font=("Segoe UI", 9), padding=(10, 6))
        style.map("Secundario.TButton", background=[("active", "#e5e7eb")])

        style.configure(
            "Accent.TButton", font=("Segoe UI", 11, "bold"), padding=(12, 10),
            background=self.COR_ACCENT, foreground="white", borderwidth=0,
        )
        style.map(
            "Accent.TButton",
            background=[("disabled", "#c7c9d9"), ("active", self.COR_ACCENT_HOVER)],
            foreground=[("disabled", "#ffffff")],
        )

        style.configure(
            "Accent.Horizontal.TProgressbar",
            background=self.COR_ACCENT, troughcolor="#e5e7eb",
            bordercolor="#e5e7eb", lightcolor=self.COR_ACCENT, darkcolor=self.COR_ACCENT,
        )

    def _criar_passo(self, parent, row, numero, titulo, variavel, texto_botao, comando):
        tk = self._tk
        ttk = self._ttk

        cartao = tk.Frame(
            parent, bg=self.COR_CARTAO, highlightthickness=1,
            highlightbackground=self.COR_BORDA, highlightcolor=self.COR_BORDA, bd=0,
        )
        cartao.grid(row=row, column=0, sticky="we", pady=(0, 12))
        cartao.grid_columnconfigure(0, weight=1)

        interno = ttk.Frame(cartao, style="Cartao.TFrame", padding=14)
        interno.grid(row=0, column=0, sticky="we")
        interno.columnconfigure(0, weight=1)

        cabecalho = ttk.Frame(interno, style="Cartao.TFrame")
        cabecalho.grid(row=0, column=0, sticky="we")
        ttk.Label(cabecalho, text=numero, style="Numero.TLabel").pack(side="left", padx=(0, 8))
        ttk.Label(
            cabecalho, text=titulo, style="CartaoTitulo.TLabel",
            wraplength=self.LARGURA_CONTEUDO - 40,
        ).pack(side="left", fill="x")

        linha = ttk.Frame(interno, style="Cartao.TFrame")
        linha.grid(row=1, column=0, sticky="we", pady=(10, 0))
        linha.columnconfigure(0, weight=1)

        ttk.Entry(linha, textvariable=variavel, state="readonly", style="Cartao.TEntry").grid(
            row=0, column=0, sticky="we"
        )
        ttk.Button(
            linha, text=texto_botao, command=comando, style="Secundario.TButton", cursor="hand2",
        ).grid(row=0, column=1, padx=(10, 0))

    def _criar_passo_capa(self, parent, row):
        """Cria o cartão do passo ④: tipo de serviço, contrato e data da capa."""
        tk = self._tk
        ttk = self._ttk

        cartao = tk.Frame(
            parent, bg=self.COR_CARTAO, highlightthickness=1,
            highlightbackground=self.COR_BORDA, highlightcolor=self.COR_BORDA, bd=0,
        )
        cartao.grid(row=row, column=0, sticky="we", pady=(0, 12))
        cartao.grid_columnconfigure(0, weight=1)

        interno = ttk.Frame(cartao, style="Cartao.TFrame", padding=14)
        interno.grid(row=0, column=0, sticky="we")
        interno.columnconfigure(0, weight=1)

        cabecalho = ttk.Frame(interno, style="Cartao.TFrame")
        cabecalho.grid(row=0, column=0, sticky="we")
        ttk.Label(cabecalho, text="④", style="Numero.TLabel").pack(side="left", padx=(0, 8))
        ttk.Label(
            cabecalho, text="Tipo de relatório, contrato e período (para a capa)",
            style="CartaoTitulo.TLabel", wraplength=self.LARGURA_CONTEUDO - 40,
        ).pack(side="left", fill="x")

        linha_tipo = ttk.Frame(interno, style="Cartao.TFrame")
        linha_tipo.grid(row=1, column=0, sticky="we", pady=(10, 0))
        ttk.Radiobutton(
            linha_tipo, text="Implantação", variable=self.tipo_var, value="implantacao",
            command=self._on_tipo_contrato_mudou,
        ).pack(side="left")
        ttk.Radiobutton(
            linha_tipo, text="Manutenção", variable=self.tipo_var, value="manutencao",
            command=self._on_tipo_contrato_mudou,
        ).pack(side="left", padx=(16, 0))
        ttk.Radiobutton(
            linha_tipo, text="Modernização", variable=self.tipo_var, value="modernizacao",
            command=self._on_tipo_contrato_mudou,
        ).pack(side="left", padx=(16, 0))

        linha_contrato = ttk.Frame(interno, style="Cartao.TFrame")
        linha_contrato.grid(row=2, column=0, sticky="we", pady=(10, 0))
        ttk.Label(linha_contrato, text="Contrato:", style="Cartao.TLabel").pack(side="left")
        combo_contrato = ttk.Combobox(
            linha_contrato, textvariable=self.contrato_label_var, values=self._labels_contrato,
            state="readonly", width=28,
        )
        combo_contrato.pack(side="left", padx=(8, 0))
        combo_contrato.bind("<<ComboboxSelected>>", lambda e: self._on_tipo_contrato_mudou())

        # Período (a maioria dos contratos) — "PERÍODO: DE ... À ..."
        self.frame_periodo = ttk.Frame(interno, style="Cartao.TFrame")
        ttk.Label(self.frame_periodo, text="De:", style="Cartao.TLabel").pack(side="left")
        ttk.Entry(
            self.frame_periodo, textvariable=self.periodo_de_var, style="Cartao.TEntry", width=12,
        ).pack(side="left", padx=(6, 14))
        ttk.Label(self.frame_periodo, text="Até:", style="Cartao.TLabel").pack(side="left")
        ttk.Entry(
            self.frame_periodo, textvariable=self.periodo_ate_var, style="Cartao.TEntry", width=12,
        ).pack(side="left", padx=(6, 0))
        ttk.Label(
            self.frame_periodo, text="(DD/MM/AAAA)", style="Status.TLabel",
        ).pack(side="left", padx=(10, 0))

        # Mês/Ano (só manutenção de Campos dos Goytacazes) — "MÊS / ANO"
        self.frame_mesano = ttk.Frame(interno, style="Cartao.TFrame")
        ttk.Label(self.frame_mesano, text="Mês:", style="Cartao.TLabel").pack(side="left")
        ttk.Combobox(
            self.frame_mesano, textvariable=self.mes_var, values=MESES_PT,
            state="readonly", width=12,
        ).pack(side="left", padx=(6, 14))
        ttk.Label(self.frame_mesano, text="Ano:", style="Cartao.TLabel").pack(side="left")
        ttk.Entry(
            self.frame_mesano, textvariable=self.ano_var, style="Cartao.TEntry", width=8,
        ).pack(side="left", padx=(6, 0))

        # Casimiro de Abreu não tem campo de data na capa (layout próprio).
        self.frame_sem_data = ttk.Frame(interno, style="Cartao.TFrame")
        ttk.Label(
            self.frame_sem_data, text="Este contrato não usa data na capa.",
            style="Status.TLabel",
        ).pack(side="left")

        for frame in (self.frame_periodo, self.frame_mesano, self.frame_sem_data):
            frame.grid(row=3, column=0, sticky="we", pady=(10, 0))

        for var in (self.periodo_de_var, self.periodo_ate_var, self.ano_var):
            var.trace_add("write", lambda *_: self._atualizar_estado())

        self._on_tipo_contrato_mudou()

    def _contrato_id_selecionado(self):
        return self._contrato_por_label[self.contrato_label_var.get()]

    def _on_tipo_contrato_mudou(self):
        """Mostra o campo de data certo (Período, Mês/Ano ou nenhum) para a
        combinação de contrato + tipo escolhida, e atualiza o botão Gerar."""
        contrato_id = self._contrato_id_selecionado()
        self.frame_periodo.grid_remove()
        self.frame_mesano.grid_remove()
        self.frame_sem_data.grid_remove()
        if usa_layout_casimiro(contrato_id):
            self.frame_sem_data.grid()
        elif usa_mes_ano(contrato_id, self.tipo_var.get()):
            self.frame_mesano.grid()
        else:
            self.frame_periodo.grid()
        self._atualizar_estado()

    def _capa_valida(self):
        """True se os campos de data necessários para a capa estão preenchidos."""
        contrato_id = self._contrato_id_selecionado()
        if usa_layout_casimiro(contrato_id):
            return True
        if usa_mes_ano(contrato_id, self.tipo_var.get()):
            return bool(self.mes_var.get()) and bool(self.ano_var.get().strip())
        return bool(self.periodo_de_var.get().strip()) and bool(self.periodo_ate_var.get().strip())

    def _montar_capa(self):
        """Monta o dict de capa {contrato_id, tipo_id, texto_data} a partir da UI."""
        contrato_id = self._contrato_id_selecionado()
        tipo_id = self.tipo_var.get()
        if usa_layout_casimiro(contrato_id):
            texto_data = ""
        elif usa_mes_ano(contrato_id, tipo_id):
            texto_data = f"{self.mes_var.get().upper()} / {self.ano_var.get().strip()}"
        else:
            rotulo = rotulo_periodo(contrato_id)
            texto_data = f"{rotulo}: {self.periodo_de_var.get().strip()} À {self.periodo_ate_var.get().strip()}"
        return {"contrato_id": contrato_id, "tipo_id": tipo_id, "texto_data": texto_data}

    def _centralizar_janela(self):
        self.root.update_idletasks()
        largura = self.root.winfo_width()
        altura = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() - largura) // 2
        y = (self.root.winfo_screenheight() - altura) // 3
        self.root.geometry(f"+{x}+{y}")

    def selecionar_planilha(self):
        from tkinter import filedialog

        caminho = filedialog.askopenfilename(
            title="Selecione o arquivo com a lista de atendimentos",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos os arquivos", "*.*")],
        )
        if caminho:
            self.planilha_var.set(caminho)
            self._atualizar_estado()

    def selecionar_fotos(self):
        from tkinter import filedialog

        caminho = filedialog.askdirectory(title="Selecione a pasta com o pack de fotos")
        if caminho:
            self.fotos_var.set(caminho)
            self._atualizar_estado()

    def selecionar_saida(self):
        from tkinter import filedialog

        caminho = filedialog.askdirectory(title="Selecione a pasta onde o PDF será salvo")
        if caminho:
            self.saida_var.set(caminho)

    def _atualizar_estado(self):
        if not hasattr(self, "gerar_btn"):
            return  # ainda montando a janela (passo da capa dispara isso antes do botão existir)
        pronto = (
            bool(self.planilha_var.get()) and bool(self.fotos_var.get()) and self._capa_valida()
        )
        self.gerar_btn.config(state="normal" if pronto else "disabled")
        if pronto:
            self.status_var.set('Tudo pronto. Clique em "Gerar Relatório".')
            self.status_label.configure(foreground=self.COR_ACCENT)
        else:
            self.status_label.configure(foreground=self.COR_TEXTO_SEC)

    def gerar(self):
        from tkinter import messagebox

        planilha_path = self.planilha_var.get()
        fotos_dir = self.fotos_var.get()

        if not os.path.isdir(fotos_dir):
            messagebox.showerror("Pasta de fotos inválida", f"Pasta não encontrada:\n{fotos_dir}")
            return

        saida_pdf = gerar_nome_saida(planilha_path, fotos_dir, self.saida_var.get() or None)
        capa = self._montar_capa()

        self.gerar_btn.config(state="disabled")
        self.status_var.set("Lendo planilha de atendimentos...")
        self.status_label.configure(foreground=self.COR_TEXTO_SEC)
        self.progress.config(mode="indeterminate")
        self.progress.start(10)

        import threading

        threading.Thread(
            target=self._gerar_em_segundo_plano, args=(planilha_path, fotos_dir, saida_pdf, capa), daemon=True
        ).start()

    def _gerar_em_segundo_plano(self, planilha_path, fotos_dir, saida_pdf, capa):
        try:
            resultado = gerar_relatorio(
                planilha_path, fotos_dir, saida_pdf, capa=capa,
                ao_avancar=lambda i, t: self.root.after(0, self._atualizar_progresso, i, t),
            )
        except Exception as e:
            self.root.after(0, self._erro, str(e))
            return
        self.root.after(0, self._concluido, saida_pdf, resultado)

    def _atualizar_progresso(self, indice, total):
        if self.progress["mode"] != "determinate":
            self.progress.stop()
            self.progress.config(mode="determinate", maximum=total, value=0)
        self.progress.config(value=indice)
        self.status_var.set(f"Gerando relatório... {indice}/{total} atendimentos")

    def _erro(self, mensagem):
        from tkinter import messagebox

        self.progress.stop()
        self.progress.config(mode="determinate", value=0)
        messagebox.showerror("Erro ao gerar relatório", mensagem)
        self.status_var.set("Ocorreu um erro. Corrija e tente novamente.")
        self.status_label.configure(foreground=self.COR_ERRO)
        self.gerar_btn.config(state="normal")

    def _concluido(self, saida_pdf, resultado):
        from tkinter import messagebox

        self.progress.config(mode="determinate", value=0)

        avisos = ""
        tem_problema = resultado.sem_foto or resultado.pastas_sem_atendimento
        if tem_problema:
            avisos = (
                "\n\nAtenção:\n"
                f"  Atendimentos sem foto: {resultado.sem_foto}\n"
                f"  Pastas de fotos sem atendimento correspondente: {len(resultado.pastas_sem_atendimento)}\n\n"
                f"Detalhes completos (números, endereços e arquivos) no log:\n{resultado.log_path}"
            )

        messagebox.showinfo(
            "Relatório gerado com sucesso",
            f"PDF salvo em:\n{saida_pdf}\n\n"
            f"Atendimentos: {resultado.total}\n"
            f"Com foto: {resultado.com_foto}\n"
            f"Sem foto: {resultado.sem_foto}"
            f"{avisos}",
        )

        self._reiniciar_formulario()

    def _reiniciar_formulario(self):
        """Limpa planilha, pasta de fotos, pasta de saída e o período para o
        próximo relatório (contrato, tipo, mês e ano ficam como estavam, já
        que normalmente se repetem de um relatório para o outro)."""
        self.planilha_var.set("")
        self.fotos_var.set("")
        self.saida_var.set("")
        self.periodo_de_var.set("")
        self.periodo_ate_var.set("")
        self.status_var.set("Selecione a planilha e a pasta de fotos para começar.")
        self._atualizar_estado()


def _habilitar_dpi_awareness():
    """No Windows, avisa o sistema que o programa lida com DPI sozinho.
    Sem isso, o Windows renderiza a janela em baixa resolução e faz upscale
    via bitmap, deixando tudo borrado em telas com escala >100%."""
    if sys.platform != "win32":
        return
    import ctypes

    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)  # system DPI aware
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def abrir_janela(planilha_inicial=None, fotos_inicial=None):
    from tkinter import Tk

    _habilitar_dpi_awareness()
    root = Tk()
    try:
        # Ajusta o fator de escala interno do Tk ao DPI real da tela,
        # já que a partir daqui o Windows não escala mais a janela por fora.
        escala = root.winfo_fpixels("1i") / 72.0
        root.tk.call("tk", "scaling", escala)
    except Exception:
        pass
    JanelaRelatorio(root, planilha_inicial, fotos_inicial)
    root.mainloop()


def parse_args():
    parser = argparse.ArgumentParser(description="Gera relatório fotográfico em PDF.")
    parser.add_argument("--planilha", dest="planilha_path", help="Caminho da planilha de atendimentos (.xlsx)")
    parser.add_argument("--fotos", dest="fotos_dir", help="Caminho da pasta com o pack de fotos")
    parser.add_argument("--saida", dest="saida_pdf", help="Caminho do PDF de saída (opcional)")
    parser.add_argument("--contrato", choices=sorted(CONTRATOS), help="Contrato do relatório (para a capa)")
    parser.add_argument("--tipo", choices=sorted(TIPOS_RELATORIO), help="Tipo de serviço (para a capa)")
    parser.add_argument("--periodo-de", dest="periodo_de", help="Início do período (DD/MM/AAAA)")
    parser.add_argument("--periodo-ate", dest="periodo_ate", help="Fim do período (DD/MM/AAAA)")
    parser.add_argument("--mes", help="Mês de referência (para contratos que usam MÊS/ANO na capa)")
    parser.add_argument("--ano", help="Ano de referência (para contratos que usam MÊS/ANO na capa)")
    return parser.parse_args()


def montar_capa_cli(args):
    """Monta o dict de capa a partir dos argumentos de linha de comando,
    ou None se --contrato/--tipo não foram informados."""
    if not args.contrato or not args.tipo:
        return None

    if usa_layout_casimiro(args.contrato):
        texto_data = ""
    elif usa_mes_ano(args.contrato, args.tipo):
        if not args.mes or not args.ano:
            sys.exit("Este contrato/tipo usa capa com MÊS/ANO: informe --mes e --ano.")
        texto_data = f"{args.mes.upper()} / {args.ano}"
    else:
        if not args.periodo_de or not args.periodo_ate:
            sys.exit("Este contrato/tipo usa capa com PERÍODO: informe --periodo-de e --periodo-ate.")
        texto_data = f"{rotulo_periodo(args.contrato)}: {args.periodo_de} À {args.periodo_ate}"

    return {"contrato_id": args.contrato, "tipo_id": args.tipo, "texto_data": texto_data}


def main():
    args = parse_args()

    # Planilha e pasta de fotos informadas por linha de comando: roda sem janela.
    if args.planilha_path and args.fotos_dir:
        executar_cli(args.planilha_path, args.fotos_dir, args.saida_pdf, capa=montar_capa_cli(args))
        return

    # Caso contrário, abre a janela guiada (uso do dia a dia).
    abrir_janela(args.planilha_path, args.fotos_dir)


if __name__ == "__main__":
    main()
